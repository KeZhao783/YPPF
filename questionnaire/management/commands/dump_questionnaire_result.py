from openpyxl import Workbook
from django.core.management.base import BaseCommand, CommandParser
from tqdm import tqdm

from questionnaire.models import Survey, AnswerSheet
import os


class Command(BaseCommand):
    help = 'Dumps the result of a questionnaire to raw_data/result.xlsx'

    def add_arguments(self, parser: CommandParser) -> None:
        parser.add_argument('questionnaire_title', type=str,
                            help='Title of questionnaire to dump')
        parser.add_argument('output_file', type=str, default='raw_data/result.xlsx', help='Output file path')
        return super().add_arguments(parser)

    def handle(self, *args, **options):
        wb = Workbook()
        ws = wb.active
        ws.title = "Result"

        survey = Survey.objects.get(title=options['questionnaire_title'])
        questions = survey.questions.order_by('order').all()

        # 若没有该文件，自动创建
        if not os.path.exists(options['output_file']):
            os.makedirs(os.path.dirname(options['output_file']), exist_ok=True)
        self.stdout.write(self.style.NOTICE(f'Survey title: {survey.title}'))

        # Add header row
        headers = [question.topic for question in questions]
        for col_num, column_title in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=col_num)
            col_letter.value = column_title

        answer_sheets = AnswerSheet.objects.filter(survey=survey)

        # Iterate through the AnswerSheet objects
        for row_num, answer_sheet in tqdm(enumerate(answer_sheets, 2), total=answer_sheets.count(), desc='Processing answer sheets'):
            answers = {
                answer.question.id: answer for answer in answer_sheet.answertext_set.all()}
            for col_num, question in enumerate(questions, 1):
                col_letter = ws.cell(row=row_num, column=col_num)
                answer = answers.get(question.id)
                if answer:
                    if question.type == 'TEXT':
                        col_letter.value = answer.body
                    elif question.type == 'SINGLE':
                        t = list(question.choices.filter(order=int(answer.body)))
                        if len(t) != 1:
                            raise ValueError(t)
                        col_letter.value = question.choices.get(
                            order=int(answer.body)).text
                    elif question.type == 'MULTIPLE':
                        choices_orders = answer.body.split(',')
                        choices_texts = [question.choices.get(
                            order=int(order)).text for order in choices_orders]
                        col_letter.value = ', '.join(choices_texts)
                else:
                    # If no answer, set as empty or a default value
                    col_letter.value = None  # or you can use "" or "N/A"

        # Save and return the workbook
        with open(options['output_file'], 'wb') as f:
            wb.save(f)

        self.stdout.write(self.style.SUCCESS(f'Results dumped to {options["output_file"]}'))